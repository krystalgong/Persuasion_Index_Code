"""
Persuasion / Style Feature Scorer (clean, modular)

Usage:
    scores = score_all(text)
    # scores is a nested dict:
    # {
    #   'Evidence': {..., 'mean': float},
    #   'Specificity': {..., 'mean': float},
    #   ...
    #   'Style': {..., 'mean': float}
    # }
"""

import os
import re
import math
import json
import logging
from pathlib import Path
from functools import lru_cache
from typing import Dict, Optional, Set, Any
import csv

# -----------------------------------------------------------------------------
# Logging to stderr only (never stdout).
# -----------------------------------------------------------------------------
logger = logging.getLogger("PI_score_generator")

# -----------------------------------------------------------------------------
# Robust paths (relative to this file, not cwd).
# Allow override via PI_HELPER_DIR for flexible deployments.
# -----------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_HELPER_DIR = BASE_DIR / "helper_features"
HELPER_DIR = Path(os.environ.get("PI_HELPER_DIR", str(DEFAULT_HELPER_DIR))).resolve()

# =========================================================
# 0) HELPERS & GLOBALS
# =========================================================

def _tok_count(text: str) -> int:
    return max(1, len(re.findall(r"\w+|\S", text)))

def _simple_tokenize(text: str):
    return re.findall(r"\b\w+\b", text.lower())

def _density(count: float, tokens: int, per: float = 100.0, k: float = 0.5) -> float:
    """Saturating map in [0,1] that is monotonic in rate per `per` tokens."""
    if tokens <= 0:
        tokens = 1
    x = (float(count) / float(tokens)) * float(per)
    return 1.0 - math.exp(-k * x)

def _optional_doc(nlp: Any, text: str):
    try:
        return nlp(text) if nlp is not None else None
    except Exception:
        return None

# -----------------------------------------------------------------------------
# Lexicon loading and cached regex compilation
# -----------------------------------------------------------------------------

# Update the cache size so it can store both the original and expanded versions if needed
@lru_cache(maxsize=2)
def _load_lexicons() -> Dict[str, Any]:
    """
    Loads lexicons.json and normalizes:
      - list -> set
      - dict[str, list] -> dict[str, set]
    """
    filename = os.environ.get("PI_LEXICON_FILE", "helper_features/lexicons.json")
    path = Path(filename)
    print(f"Loading lexicons from: {filename}")
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    norm: Dict[str, Any] = {}
    for k, v in data.items():
        if isinstance(v, list):
            norm[k] = set(v)
        elif isinstance(v, dict):
            vv = {}
            for kk, ll in v.items():
                vv[kk] = set(ll) if isinstance(ll, list) else ll
            norm[k] = vv
        else:
            norm[k] = v
    return norm

_PATTERN_CACHE: Dict[str, re.Pattern] = {}

def _compile_vocab_pattern(vocab: Set[str]) -> re.Pattern:
    """
    Compiles a case-insensitive pattern that matches any phrase in vocab.
    Uses word-ish boundaries that work for multi-word phrases too.
    """
    items = sorted((s for s in vocab if isinstance(s, str) and s.strip()), key=len, reverse=True)
    if not items:
        return re.compile(r"$.")

    parts = [re.escape(s.strip()) for s in items]
    pattern = r"(?<!\w)(?:" + "|".join(parts) + r")(?!\w)"
    try:
        return re.compile(pattern, re.IGNORECASE)
    except re.error:
        logger.warning("Failed to compile vocab regex (size=%d). Falling back to no matches.", len(items))
        return re.compile(r"$.")

# def _count_vocab(text: str, cache_key: str, vocab: Set[str]) -> int:
#     pat = _PATTERN_CACHE.get(cache_key)
#     if pat is None:
#         pat = _compile_vocab_pattern(vocab)
#         _PATTERN_CACHE[cache_key] = pat
#     return len(pat.findall(text))

def _count_vocab(text: str, cache_key: str, vocab: Set[str]) -> int:
    pat = _PATTERN_CACHE.get(cache_key)
    if pat is None:
        pat = _compile_vocab_pattern(vocab)
        _PATTERN_CACHE[cache_key] = pat

    # Collect all match spans, then remove any span fully contained within a longer match.
    # e.g. "besides" is absorbed by "besides that" — only the longer match counts.
    matches = [(m.start(), m.end()) for m in pat.finditer(text)]
    non_overlapping = [
        (s, e) for (s, e) in matches
        if not any(s2 <= s and e <= e2 and (s2, e2) != (s, e) for (s2, e2) in matches)
    ]
    return len(non_overlapping)

def _count_lex(text: str, lex_key: str) -> int:
    lex = _load_lexicons().get(lex_key)
    if not isinstance(lex, set) or not lex:
        return 0
    return _count_vocab(text, lex_key, lex)

# -----------------------------------------------------------------------------
# Regexes
# -----------------------------------------------------------------------------
RE_NUMBER = re.compile(r"\b\d[\d,.\%]*\b")
RE_PERCENT = re.compile(r"\b\d+%\b|\b\d+\spercent\b", re.IGNORECASE)
RE_URL = re.compile(r"https?://\S+|\bwww\.\S+", re.IGNORECASE)
RE_APA = re.compile(r"\b[A-Z][a-z]+(?:\s[A-Z][a-z]+)?\s*\(\d{4}\)")
RE_NUM_CIT = re.compile(r"\[(?:\d{1,3}(?:,\s*\d{1,3})*)\]")

RE_TIME_SPECIFIC = re.compile(r"\b(\d{1,2}:\d{2}|o'clock|am|pm|morning|afternoon|evening|weekend)\b", re.IGNORECASE)
RE_DATE_MD = re.compile(r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\.?\s+\d{1,2}(?:,\s*\d{2,4})?\b", re.IGNORECASE)

RE_TITLES = re.compile(r"\b(dr\.|prof\.|professor|md|m\.d\.|phd|ph\.d\.)\b", re.IGNORECASE)
RE_AUTH_ORG = re.compile(
    r"\b("
    r"world health organization|who|harvard university|cdc|nih|fda|united nations|university|"
    r"institute|organization|agency|center|centre|department|researchers at|study by"
    r")\b",
    re.IGNORECASE,
)
# RE_CONSENSUS = re.compile(
#     r"\b(\d{1,3}%|majority|most|almost all|nearly everyone|millions|thousands)\b\s+(of\s+people|agree|support|believe|concur)\b",
#     re.IGNORECASE,
# )
RE_CONSENSUS = re.compile(
    r"("
    r"\b(?:over|about|around|approximately|nearly|almost)?\s*\d{1,3}%\s+of\s+\w+"
    r"|"
    r"\b(?:over|about|around|approximately|nearly|almost)?\s*\d{1,3}\s+percent\s+of\s+\w+"
    r"|"
    r"\bmajority of\b"
    r"|"
    r"\bmost people\b"
    r"|"
    r"\bmost adults\b"
    r"|"
    r"\bmost participants\b"
    r"|"
    r"\balmost all\b"
    r"|"
    r"\bnearly everyone\b"
    r")",
    re.IGNORECASE,
)

RE_QUESTION = re.compile(r"\?")
WORD_RE = re.compile(r"[A-Za-z]+(?:'[A-Za-z]+)?")

RECIPROCITY_PATTERNS = [
    r"i (will|'ll) (pay|give|return) (it|you) (back|forward)",
    r"(return|reciprocate) the favor",
    r"i (will|'ll) do the same for you",
    r"in (return|exchange),? i (will|'ll)",
    r"i (will|'ll) (award|give) (you )?a (delta|point|upvote)",
    r"you (will|'ll) have my (thanks|gratitude|respect)",
    r"i (will|'ll) (credit|cite) you",
    r"i (will|'ll) (change|update|reconsider) my (view|mind|stance)",
    r"i (will|'ll) look into (that|your suggestion)",
    r"help (others|someone else) (later|down the line|in the future)",
    r"pass (this|it) (on|along)",
]

RE_SOFT_CONCESSION = re.compile(
    r"\b("
    r"would you be open to|just wanted to|might be helpful|slightly different|"
    r"i would appreciate|would it help|perhaps we could|maybe we could"
    r")\b",
    re.IGNORECASE,
)

RE_TIME_SPAN = re.compile(
    r"\b(within|in|over|after)\s+"
    r"(\d+|one|two|three|four|five|six|seven|eight|nine|ten)\s+"
    r"(day|days|week|weeks|month|months|year|years)\b",
    re.IGNORECASE,
)

RE_FUTURE_FRAME = re.compile(
    r"\b("
    r"future generations|future|next generation|imagine a world|imagine|"
    r"before it is too late|too late|if we fail to act|if we do not act|"
    r"if we don't act|no longer guaranteed|becomes a luxury"
    r")\b",
    re.IGNORECASE,
)

RE_SOFT_GAIN = re.compile(
    r"\b(better solution|better approach|better strategy|might be helpful|helpful)\b",
    re.IGNORECASE,
)

RE_WE = re.compile(r"\b(we|our|us)\b", re.IGNORECASE)
RE_WE_care = re.compile(r"\b(we care about|our future|our children|our world)\b", re.IGNORECASE)

RE_IMAGERY = re.compile(
    r"\b(imagine|picture|visualize|a world where|see a future|envision)\b",
    re.IGNORECASE,
)

RE_SOFT_ACK = re.compile(
    r"\b("
    r"current approach|existing approach|limitations with the current approach|"
    r"we have already seen|some limitations|the current strategy"
    r")\b",
    re.IGNORECASE,
)

# -----------------------------------------------------------------------------
# LIWC: cached loading + cached regex compilation
# -----------------------------------------------------------------------------

@lru_cache(maxsize=1)
def _load_liwc_dic() -> Dict[str, list]:
    path = HELPER_DIR / "en_liwc.txt"
    dic: Dict[str, list] = {}
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if ": " not in line:
                continue
            key, rest = line.split(": ", 1)
            toks = [t.strip() for t in rest.split() if t.strip()]
            dic[key] = toks
    return dic

_LIWC_REGEX_CACHE: Dict[str, re.Pattern] = {}

def _compile_liwc_regex(category_name: str) -> re.Pattern:
    dic = _load_liwc_dic()
    patterns = dic.get(category_name, [])
    if not patterns:
        return re.compile(r"$.")

    regex_parts = []
    for p in patterns:
        p = p.strip().lower()
        if not p:
            continue
        if p.endswith("*"):
            clean = re.escape(p[:-1])
            regex_parts.append(rf"\b{clean}\w*")
        else:
            regex_parts.append(rf"\b{re.escape(p)}\b")

    if not regex_parts:
        return re.compile(r"$.")

    try:
        return re.compile("|".join(regex_parts), re.IGNORECASE)
    except re.error:
        logger.warning("Failed to compile LIWC regex for category '%s'.", category_name)
        return re.compile(r"$.")

def _liwc_re(category: str) -> re.Pattern:
    pat = _LIWC_REGEX_CACHE.get(category)
    if pat is None:
        pat = _compile_liwc_regex(category)
        _LIWC_REGEX_CACHE[category] = pat
    return pat

# =========================================================
# Lazy spaCy + VADER loaders (cached per process)
# =========================================================

@lru_cache(maxsize=1)
def _get_nlp():
    """
    spaCy is heavy. Load once per worker process.
    You can disable spaCy loading by setting PI_DISABLE_SPACY=1.
    """
    if os.environ.get("PI_DISABLE_SPACY", "").strip() == "1":
        return None
    try:
        import spacy  # imported lazily
        return spacy.load("en_core_web_sm")
    except Exception as e:
        logger.warning("spaCy model unavailable (en_core_web_sm). NER-based features will be 0.0. Error: %s", e)
        return None

@lru_cache(maxsize=1)
def _get_vader():
    try:
        from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
        return SentimentIntensityAnalyzer()
    except Exception as e:
        logger.warning("VADER unavailable. vader_compound will be 0.0. Error: %s", e)
        return None

# =========================================================
# Concreteness dataset loader (cached per process)
# =========================================================

@lru_cache(maxsize=1)
def _load_concreteness_dic() -> Dict[str, float]:
    """
    Load Brysbaert concreteness dataset once per worker process.
    Returns {} if missing/unreadable.
    Uses openpyxl read_only mode for low memory overhead.
    """
    path = HELPER_DIR / "Brysbaert_concretness_dataset.xlsx"
    if not path.exists():
        return {}

    try:
        from openpyxl import load_workbook  # lazy import
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return {}

        header_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
        if "Word" not in header_map or "Conc.M" not in header_map:
            return {}

        w_idx = header_map["Word"]
        c_idx = header_map["Conc.M"]

        out: Dict[str, float] = {}
        for r in rows:
            if not r:
                continue
            if w_idx >= len(r) or c_idx >= len(r):
                continue
            w = r[w_idx]
            c = r[c_idx]
            if w is None or c is None:
                continue
            ws_ = str(w).strip().lower()
            if not ws_:
                continue
            try:
                out[ws_] = float(c)
            except Exception:
                continue

        return out
    except Exception as e:
        logger.warning("Concreteness dataset unavailable. lexical_concreteness will use baseline. Error: %s", e)
        return {}
    
@lru_cache(maxsize=1)
def _load_mwe_concreteness_dic(path="helper_features/MultiwordExpression_Concreteness_Ratings.csv"):
    mwe_dic = {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                expr = row["Expression"].strip().lower()
                score = row["Mean_C"]

                if score == "NA" or score == "":
                    continue

                mwe_dic[expr] = float(score)
    except Exception:
        return {}

    return mwe_dic
    
# load NRC-VAD lexicons
@lru_cache(maxsize=1)
def _load_nrc_vad(path="helper_features/NRC-VAD-Lexicon-v2.1/Unigrams/unigrams-NRC-VAD-Lexicon-v2.1.txt"):
    vad = {}
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            term = row["term"].lower()
            vad[term] = (
                float(row["valence"]),
                float(row["arousal"]),
                float(row["dominance"]),
            )
    return vad

# =========================================================
# 1) EVIDENCE
# =========================================================

def evidence_statistical(text: str) -> float:
    num_cnt = len(RE_NUMBER.findall(text))
    per_cnt = len(RE_PERCENT.findall(text))
    unit_cnt = _count_lex(text, "EVI_UNITS")
    return _density(num_cnt + per_cnt + unit_cnt, _tok_count(text))

def evidence_attribution(text: str) -> float:
    c = len(RE_APA.findall(text)) + len(RE_NUM_CIT.findall(text)) + len(RE_URL.findall(text))
    c += _count_lex(text, "AUTHORITY_PHRASE")
    d = _density(c, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def evidence_named_entities(text: str, nlp=None) -> float:
    doc = _optional_doc(nlp, text)
    if doc is None:
        return 0.0
    keep = {"PERSON", "ORG", "GPE", "DATE", "QUANTITY", "CARDINAL", "MONEY", "PERCENT", "NORP"}
    ner_cnt = sum(1 for e in doc.ents if e.label_ in keep)
    return _density(ner_cnt, max(1, len(doc)))

# =========================================================
# 2) LOGIC / COHESION
# =========================================================

def logic_structural_reasoning(text: str) -> float:
    cnt = _count_lex(text, "LOGIC_CAUSAL")
    cnt += _count_lex(text, "LOGIC_INFERENCE")
    cnt += _count_lex(text, "LOGIC_REFERENCE")
    if_then = len(re.findall(r"\bif\b.*\bthen\b", text, re.IGNORECASE | re.DOTALL))
    return _density(cnt + if_then, _tok_count(text))

def logic_discourse_cohesion(text: str) -> float:
    count = _count_lex(text, "LOGIC_CONTRAST")
    count += _count_lex(text, "LOGIC_ADDITIVE")
    # We count how many non-trivial words are repeated. 
    # High repetition often indicates "staying on message" in rhetoric.
    words = re.findall(r'\b\w{4,}\b', text.lower()) # Only words with 4+ letters
    if len(words) > 0:
        unique_words = set(words)
        repetition_count = len(words) - len(unique_words)
    else:
        repetition_count = 0
    return _density(count + repetition_count, _tok_count(text))


# =========================================================
# 3) ARGUMENTATION
# =========================================================

def argument_conclusion_explicitness(text: str) -> float:
    count = _count_lex(text, "ARG_CLAIM")
    count += _count_lex(text, "ARG_CONCLUSION")
    return _density(count, _tok_count(text))

def argument_premise_density(text: str) -> float:
    d = _density(_count_lex(text, "ARG_PREMISE"), _tok_count(text))
    return 1.0 if d > 0 else 0.0

def argument_quantity_intensity(text: str) -> float:
    """
    Calculates the 'Argumentative Intensity' of a text.
    Reflects the 'Length is Strength' heuristic (Petty & Cacioppo, 1984).
    
    Formula: Argumentative Sentences / Total Sentences
    A sentence is argumentative if it contains at least one claim, conclusion, or premise marker.
    """
    sentences = [s.strip() for s in re.split(r'[.!?]+', text) if s.strip()]
    sent_count = len(sentences)

    if sent_count == 0:
        return 0.0

    argumentative = sum(
        1 for s in sentences
        if _count_lex(s, "ARG_CLAIM") > 0
        or _count_lex(s, "ARG_CONCLUSION") > 0
        or _count_lex(s, "ARG_PREMISE") > 0
    )

    return argumentative / sent_count

# def argument_style_sophistication(text: str) -> float:
#     toks = _tok_count(text)
#     lex = _load_lexicons().get("ARG_STYLE", {})
#     if not isinstance(lex, dict):
#         return 0.0

#     cnt = 0
#     for kk, v in lex.items():
#         if not isinstance(v, set) or not v:
#             continue
#         cache_key = f"ARG_STYLE::{kk}"
#         cnt += _count_vocab(text, cache_key, v)

#     return _density(cnt, toks)

def argument_style_sophistication(text: str) -> float:
    toks = _tok_count(text)
    lex = _load_lexicons()
    if not isinstance(lex, dict):
        return 0.0

    # Collect all ARG_STYLE sub-category keys
    arg_style_keys = {
        k: v for k, v in lex.items()
        if k.startswith("ARG_STYLE.") and isinstance(v, (set, list)) and v
    }

    if not arg_style_keys:
        return 0.0

    # Count matches per sub-category, but reward DIVERSITY of styles used
    categories_hit = 0
    total_cnt = 0

    for key, vocab in arg_style_keys.items():
        vocab_set = set(vocab) if not isinstance(vocab, set) else vocab
        cnt = _count_vocab(text, key, vocab_set)
        if cnt > 0:
            categories_hit += 1
            total_cnt += cnt

    num_categories = len(arg_style_keys)  # e.g. 6

    # Base density score (raw usage rate)
    density_score = _density(total_cnt, toks)

    # Diversity bonus: fraction of style categories used (0.0 → 1.0)
    diversity_score = categories_hit / num_categories

    # Blend: weight diversity more heavily since sophistication = variety of rhetorical moves
    raw = 0.4 * density_score + 0.6 * diversity_score

    return min(raw, 1.0)

# =========================================================
# 4) SPECIFICITY
# =========================================================

def spec_psychological_nearness(text: str) -> float:
    toks = [t.strip(".,;:!?()[]").lower() for t in text.split()]
    total_toks = len(toks) if toks else 1

    time_regex_hits = len(RE_DATE_MD.findall(text)) + len(RE_TIME_SPECIFIC.findall(text))
    near_hits = _count_lex(text, "SPEC_PSYCH_NEAR")
    far_hits = _count_lex(text, "SPEC_PSYCH_FAR")

    anchor_hits = _count_lex(text, "SPEC_ANCHORS")
    vague_penalty = _count_lex(text, "SPEC_VAGUE")

    net_count = (time_regex_hits + near_hits + anchor_hits) - (far_hits * 0.5 + vague_penalty * 0.5)
    d = _density(net_count, total_toks)
    return 1.0 if d > 0 else 0.0

def _match_mwe_scores(text: str, mwe_dic: dict):
    text_l = text.lower()
    scores = []

    for phrase, score in mwe_dic.items():
        if phrase in text_l:
            scores.append(score)

    return scores

def spec_lexical_concreteness(text: str) -> float:
    conc_dic = _load_concreteness_dic()
    mwe_dic = _load_mwe_concreteness_dic()

    if not conc_dic and not mwe_dic:
        return 0.5

    # --- Unigram scores ---
    toks = [t.strip(".,;:!?()[]").lower() for t in text.split()]
    unigram_scores = [conc_dic[t] for t in toks if t in conc_dic]

    # --- Multi-word scores ---
    mwe_scores = _match_mwe_scores(text, mwe_dic)

    scores = unigram_scores + [min(s + math.log1p(s), 5) for s in mwe_scores]

    if not scores:
        return 0.5

    avg = sum(scores) / len(scores)

    # Normalize (1–5 → 0–1)
    return (avg - 1.0) / 4.0

def spec_interactional_immediacy(text: str, nlp=None) -> float:
    toks = _tok_count(text)
    you_hits = len(_liwc_re("You").findall(text))

    person_near_you = 0
    doc = _optional_doc(nlp, text)
    if doc is not None:
        you_idxs = {i for i, t in enumerate(doc) if t.text.lower() in {"you", "your", "you're"}}
        for ent in doc.ents:
            if ent.label_ == "PERSON":
                center = (ent.start + ent.end) // 2
                if any(abs(center - yi) <= 6 for yi in you_idxs):
                    person_near_you += 1

    return min(1.0, (_density(you_hits, toks) * 0.7 + _density(person_near_you, toks) * 0.3) * 2.0)

# =========================================================
# 5) OPPONENT’S VIEW
# =========================================================

def opponent_acknowledge(text: str) -> float:
    count = _count_lex(text, "OPP_VIEW")
    count += len(RE_SOFT_ACK.findall(text))
    d = _density(count, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def opponent_refutation_strength(text: str) -> float:
    lex = _load_lexicons()
    opp_view = lex.get("OPP_VIEW", set())
    contrast = lex.get("LOGIC_CONTRAST", set())
    if not isinstance(opp_view, set) or not isinstance(contrast, set):
        return 0.0

    t = re.sub(r"\s+", " ", text.lower())

    idx_ack = -1
    for p in opp_view:
        idx = t.find(p)
        if idx != -1:
            idx_ack = idx
            break

    if idx_ack == -1:
        return 0.0

    has_refute = any(t.find(m, idx_ack) > idx_ack for m in contrast)
    return 1.0 if has_refute else 0.5

# =========================================================
# 6) Ethos: AUTHORITY / CREDIBILITY
# =========================================================

def authority_titles(text: str) -> float:
    return _density(len(RE_TITLES.findall(text)), _tok_count(text))

def authority_organizations(text: str, nlp=None) -> float:
    # simple regex hits
    regex_hits = len(RE_AUTH_ORG.findall(text))

    # optional spaCy ORG entities
    org_hits = 0
    doc = _optional_doc(nlp, text)
    if doc is not None:
        org_hits = sum(1 for ent in doc.ents if ent.label_ == "ORG")
    return _density(regex_hits + org_hits, _tok_count(text))

def authority_phrases(text: str) -> float:
    d = _density(_count_lex(text, "AUTHORITY_PHRASE"), _tok_count(text))
    return 1.0 if d > 0 else 0.0

# def authority_attribution(text: str, nlp=None) -> float:
#     doc = _optional_doc(nlp, text)
#     if doc is None:
#         return 0.0
#     lex = _load_lexicons()
#     verbs = lex.get("REPORTING_VERBS", set())
#     if not isinstance(verbs, set) or not verbs:
#         return 0.0

#     verb_idxs = {i for i, t in enumerate(doc) if t.lemma_.lower() in verbs}
#     pairs = 0
#     if verb_idxs:
#         for ent in doc.ents:
#             if ent.label_ in {"PERSON", "ORG"}:
#                 center = (ent.start + ent.end) // 2
#                 if any(abs(center - vi) <= 3 for vi in verb_idxs):
#                     pairs += 1
#     return _density(pairs, max(1, len(doc)))

def authority_consensus(text: str) -> float:
    matches = len(RE_CONSENSUS.findall(text))
    d = _density(matches, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def authority_speech_power(text: str) -> float:
    hedge_cnt = _count_lex(text, "HEDGES")
    hesitation_cnt = _count_lex(text, "HESITATIONS")
    tag_cnt = _count_lex(text, "AUT_TAG_QUESTION")

    total_powerless = hedge_cnt + hesitation_cnt + tag_cnt
    raw_power = 1.0 - _density(total_powerless, _tok_count(text))
    return max(0.0, raw_power)

# =========================================================
# 7) POLITENESS
# =========================================================

def politeness_professional_courtesy(text: str) -> float:
    c = _count_lex(text, "PLEASE")
    c += _count_lex(text, "THANKS")
    c += _count_lex(text, "COURTESY")

    c += _count_lex(text, "Apology")
    c += _count_lex(text, "Reassurance")
    c += _count_lex(text, "Greeting")

    d = _density(c, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def politeness_rapport_building(text: str) -> float:
    count = _count_lex(text, "RAPPORT")

    # New additions
    count += _count_lex(text, "Greeting")
    count += _count_lex(text, "Filler")
    count += _count_lex(text, "By.The.Way")
    count += _count_lex(text, "Actually")

    d = _density(count, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def politeness_non_imposition(text: str) -> float:
    count = _count_lex(text, "HEDGES") + _count_lex(text, "INDIRECTS")
    
    # New additions
    count += _count_lex(text, "Subjunctive")
    count += _count_lex(text, "Indicative")
    count += _count_lex(text, "For.Me")
    count += _count_lex(text, "For.You")

    d = _density(count, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def politeness_domineering(text: str) -> float:
    count = _count_lex(text, "DOMINEERING")
    d = _density(count, _tok_count(text))
    return 1.0 if d > 0 else 0.0

# =========================================================
# 8) COMMITMENT
# =========================================================

def commitment_statements(text: str) -> float:
    d = _density(_count_lex(text, "COMMITMENT"), _tok_count(text))
    return 1.0 if d > 0 else 0.0

def commitment_power(text: str) -> float:
    lex = _load_lexicons()
    commitment = lex.get("COMMITMENT", set())
    proof = lex.get("COMMITMENT_PROOF", set())
    if not isinstance(commitment, set) or not isinstance(proof, set):
        return 0.0

    t = re.sub(r"\s+", " ", text.lower())

    idx_commit = -1
    for p in commitment:
        idx = t.find(p)
        if idx != -1:
            idx_commit = idx
            break

    if idx_commit == -1:
        return 0.0

    has_power = any(t.find(m, idx_commit) > idx_commit for m in proof)
    return 1.0 if has_power else 0.5

# =========================================================
# 9) STYLE
# =========================================================
from wordfreq import zipf_frequency

def style_fluency(text: str) -> float:
    tokens = [w for w in WORD_RE.findall(text)]
    if not tokens:
        return 0.0
    
    def is_oov(w):
        return zipf_frequency(w, "en") < 2.5

    oov = sum(1 for w in tokens if is_oov(w))
    score = min(1.0, oov / len(tokens))
    return math.exp(-5 * score)

def style_length(text: str) -> float:
    return min(1.0, math.log1p(_tok_count(text)) / 7.0)

def style_rhetorical_punctuation(text: str) -> float:
    toks = max(1, _tok_count(text))
    quotes_score = _density(len(re.findall(r"[\"“”‘’']", text)), toks)
    questions_score = _density(text.count("?"), toks)
    excl_score = _density(text.count("!"), toks)
    return min(1.0, (quotes_score + questions_score + excl_score) / 3.0)

# =========================================================
# 10) SENTIMENT
# =========================================================

def sentiment_polarity(text: str, vader=None) -> float:
    if vader is None:
        return 0.0
    try:
        score = vader.polarity_scores(text).get("compound", 0.0)
        return min(1.0, abs(float(score)))
    except Exception:
        return 0.0

def sentiment_language_intensity(text: str) -> float:
    lex = _load_lexicons().get("SENT_INTENSITY", {})
    text = text.lower()
    tok_n = _tok_count(text)
    if tok_n == 0:
        return 0.0

    # weighted sentiment/intensifier words
    # count_extreme = _count_vocab(text, "SENT_INTENSITY::extreme", lex.get("extreme", set()))
    # count_strong  = _count_vocab(text, "SENT_INTENSITY::strong",  lex.get("strong", set()))
    # count_mod     = _count_vocab(text, "SENT_INTENSITY::moderate", lex.get("moderate", set()))
    # count_weak    = _count_vocab(text, "SENT_INTENSITY::weak",    lex.get("weak", set()))

    count_extreme = _count_lex(text, "SENT_INTENSITY.extreme")
    count_strong = _count_lex(text, "SENT_INTENSITY.strong")
    count_mod = _count_lex(text, "SENT_INTENSITY.moderate")
    count_weak = _count_lex(text, "SENT_INTENSITY.weak")

    lexical_score = (
        3.0 * count_extreme +
        2.0 * count_strong +
        1.0 * count_mod +
        0.5 * count_weak
    )

    # directive / urgency force
    directive_score = (
        1.5 * _count_lex(text, "DOMINEERING") +
        1.2 * _count_lex(text, "URGENCY_TIME") +
        1.0 * _count_lex(text, "ARG_CLAIM") +
        0.8 * _count_lex(text, "IMPACT_THREAT")
    )

    score = lexical_score + directive_score
    return min(1.0, _density(score, tok_n))

def sentiment_anger(text: str) -> float:
    matches = _liwc_re("Anger").findall(text)
    return min(1.0, _density(len(matches), _tok_count(text)))

def sentiment_sadness(text: str) -> float:
    matches = _liwc_re("Sad").findall(text)
    return min(1.0, _density(len(matches), _tok_count(text)))

def sentiment_fear_threat(text: str) -> float:
    liwc_hits = len(_liwc_re("Anx").findall(text))
    lex_hits = _count_lex(text, "PROP_FEAR") + _count_lex(text, "IMPACT_THREAT")
    return min(1.0, _density(liwc_hits + lex_hits, _tok_count(text)))

def sentiment_joy_gain(text: str) -> float:
    matches = _liwc_re("Posemo").findall(text)
    return min(1.0, _density(len(matches), _tok_count(text)))

def _normalized_vad(x):
    return (x + 1) / 2

def sentiment_vad_scores(text: str) -> dict:
    vad_lex = _load_nrc_vad()
    tokens = [t.strip(".,;:!?()[]").lower() for t in text.split()]

    v_sum, a_sum, d_sum = 0.0, 0.0, 0.0
    count = 0

    for tok in tokens:
        if tok in vad_lex:
            v, a, d = vad_lex[tok]
            v_sum += v
            a_sum += a
            d_sum += d
            count += 1

    if count == 0:
        return {"valence": 0.0, "arousal": 0.0, "dominance": 0.0}

    return {
        "valence": _normalized_vad(v_sum / count),
        "arousal": _normalized_vad(a_sum / count),
        "dominance": _normalized_vad(d_sum / count),
    }

# =========================================================
# 11) IMPACT
# =========================================================

def impact_gain_framing(text: str) -> float:
    count = _count_lex(text, "IMPACT")
    count += _count_lex(text, "IMPACT_GAIN")
    count += len(RE_SOFT_GAIN.findall(text))
    d = _density(count, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def impact_loss_framing(text: str) -> float:
    d = _density(_count_lex(text, "IMPACT_LOSS"), _tok_count(text))
    return 1.0 if d > 0 else 0.0

def impact_threat_severity(text: str) -> float:
    d = _density(_count_lex(text, "IMPACT_THREAT"), _tok_count(text))
    return 1.0 if d > 0 else 0.0

def impact_future_projection(text: str) -> float:
    future_hits = _count_lex(text, "IMPACT_WILL")
    span_hits = len(RE_TIME_SPAN.findall(text))
    frame_hits = len(RE_FUTURE_FRAME.findall(text))
    d = _density(future_hits + span_hits + frame_hits, _tok_count(text))
    return 1.0 if d > 0 else 0.0

# =========================================================
# 12) ENGAGEMENT
# =========================================================

def engagement_identification_pov(text: str) -> float:
    I_identity_hits = len(_liwc_re("I").findall(text))
    we_hits = len(RE_WE.findall(text))
    identity_hits = len(RE_WE_care.findall(text))
    return _density(I_identity_hits + we_hits + identity_hits, _tok_count(text))

def engagement_self_referencing(text: str) -> float:
    return _density(len(_liwc_re("You").findall(text)), _tok_count(text))

def engagement_inquiry(text: str) -> float:
    d = _density(len(RE_QUESTION.findall(text)), _tok_count(text))
    return 1.0 if d > 0 else 0.0

def engagement_past(text: str) -> float:
    past_hits = len(_liwc_re("Past").findall(text))
    return _density(past_hits, _tok_count(text))

def engagement_imagery(text: str) -> float:
    imagery_hits = len(_liwc_re("See").findall(text))
    regex_hits = len(RE_IMAGERY.findall(text))
    return _density(imagery_hits + regex_hits, _tok_count(text))

def engagement_characters(text: str) -> float:
    character_hits = len(_liwc_re("Ppron").findall(text))
    return _density(character_hits, _tok_count(text))

# =========================================================
# 13) RECIPROCITY
# =========================================================

def reciprocity_direct_promise(text: str) -> float:
    text_lower = text.lower()
    for pattern in RECIPROCITY_PATTERNS:
        if re.search(pattern, text_lower):
            return 1.0
    return 0.0

def reciprocity_mutual_benefit(text: str) -> float:
    d = _density(_count_lex(text, "REC_MUTUAL_EXCHANGE"), _tok_count(text))
    return 1.0 if d > 0 else 0.0

def reciprocity_concession_framing(text: str) -> float:
    count = _count_lex(text, "REC_CONCESSION")
    count += len(RE_SOFT_CONCESSION.findall(text))
    d = _density(count, _tok_count(text))
    return 1.0 if d > 0 else 0.0

# =========================================================
# 14) SCARCITY / URGENCY
# =========================================================

def scarcity_temporal_urgency(text: str) -> float:
    psych_near = _count_lex(text, "SPEC_PSYCH_NEAR")
    lex_hits = _count_lex(text, "URGENCY_TIME")
    return _density(psych_near + lex_hits, _tok_count(text))

def scarcity_exclusivity_quantity(text: str) -> float:
    return _density(_count_lex(text, "SCARCITY_QUANTITY"), _tok_count(text))

# =========================================================
# 15) PROPAGANDA CUES
# =========================================================

def prop_emotional_charge(text: str) -> float:
    cnt = _count_lex(text, "PROP_LOADED")
    cnt += _count_lex(text, "PROP_NAMECALL")
    cnt += _count_lex(text, "PROP_FEAR")
    cnt += _count_lex(text, "IMPACT_LOSS")
    cnt += _count_lex(text, "IMPACT_THREAT")
    d = _density(cnt, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def prop_logical_distortion(text: str) -> float:
    cnt = _count_lex(text, "PROP_WHATABOUT")
    cnt += _count_lex(text, "PROP_STRAW")
    cnt += _count_lex(text, "PROP_VAGUE")
    cnt += _count_lex(text, "PROP_OVERSIMPLIFY")
    d = _density(cnt, _tok_count(text))
    return 1.0 if d > 0 else 0.0

def prop_heuristic_identity_appeals(text: str) -> float:
    cnt = _count_lex(text, "PROP_FLAG")
    cnt += _count_lex(text, "PROP_BANDWAGON")
    cnt += _count_lex(text, "PROP_SLOGAN")
    cnt += _count_lex(text, "PROP_AUTH")
    return _density(cnt, _tok_count(text))



# =========================================================
# MASTER API
# =========================================================

def score_all(text: str) -> Dict[str, Dict[str, float]]:
    """
    Return nested scores per category.
    Heavy resources are loaded lazily and cached per worker process.
    """
    nlp = _get_nlp()
    vader = _get_vader()

    out: Dict[str, Dict[str, float]] = {}

    text = re.sub(r"\s+", " ", text.lower()).strip()

    out["Evidence"] = {
        "statistical": evidence_statistical(text),
        "attribution": evidence_attribution(text),
        "named_entities": evidence_named_entities(text, nlp=nlp),
    }
    out["Evidence"]["mean"] = sum(out["Evidence"].values()) / len(out["Evidence"].values())

    out["Specificity"] = {
        "psychological_nearness": spec_psychological_nearness(text),
        "lexical_concreteness": spec_lexical_concreteness(text),
        "interactional_immediacy": spec_interactional_immediacy(text, nlp=nlp),
    }
    out["Specificity"]["mean"] = sum(out["Specificity"].values()) / len(out["Specificity"].values())

    out["Authority/Credibility"] = {
        "titles": authority_titles(text),
        "organizations": authority_organizations(text, nlp=nlp),
        "phrases": authority_phrases(text),
        "consensus": authority_consensus(text),
        "speech_power": authority_speech_power(text),
    }
    out["Authority/Credibility"]["mean"] = sum(out["Authority/Credibility"].values()) / len(out["Authority/Credibility"].values())

    out["Logic/Cohesion"] = {
        "structural_reasoning": logic_structural_reasoning(text),
        "discourse_cohesion": logic_discourse_cohesion(text),
    }
    out["Logic/Cohesion"]["mean"] = sum(out["Logic/Cohesion"].values()) / len(out["Logic/Cohesion"].values())

    out["Argumentation"] = {
        "conclusion_explicitness": argument_conclusion_explicitness(text),
        "premise_density": argument_premise_density(text),
        "quantity_intensity": argument_quantity_intensity(text),
        "style_sophistication": argument_style_sophistication(text),
    }
    out["Argumentation"]["mean"] = sum(out["Argumentation"].values()) / len(out["Argumentation"].values())

    out["Opponent’s View"] = {
        "acknowledge": opponent_acknowledge(text),
        "refutation_strength": opponent_refutation_strength(text),
    }
    out["Opponent’s View"]["mean"] = sum(out["Opponent’s View"].values()) / len(out["Opponent’s View"].values())

    vad = sentiment_vad_scores(text)
    out["Sentiment"] = {
        "vader_compound": sentiment_polarity(text, vader=vader),
        "language_intensity": sentiment_language_intensity(text),
        "fear_threat": sentiment_fear_threat(text),
        "joy_gain": sentiment_joy_gain(text),
        "anger": sentiment_anger(text),
        "sadness": sentiment_sadness(text),
        "valence": vad["valence"],
        "arousal": vad["arousal"],
        "dominance": vad["dominance"]
    }
    out["Sentiment"]["mean"] = sum(out["Sentiment"].values()) / len(out["Sentiment"].values())

    out["Politeness"] = {
        "professional_courtesy": politeness_professional_courtesy(text),
        "rapport_building": politeness_rapport_building(text),
        "non_imposition": politeness_non_imposition(text),
        "domineering": politeness_domineering(text)
    }
    out["Politeness"]["mean"] = sum(out["Politeness"].values()) / len(out["Politeness"].values())

    out["Reciprocity"] = {
        "direct_promise": reciprocity_direct_promise(text),
        "mutual_benefit": reciprocity_mutual_benefit(text),
        "concession_framing": reciprocity_concession_framing(text),
    }
    out["Reciprocity"]["mean"] = sum(out["Reciprocity"].values()) / len(out["Reciprocity"].values())

    out["Impact"] = {
        "gain_framing": impact_gain_framing(text),
        "loss_framing": impact_loss_framing(text),
        "threat_severity": impact_threat_severity(text),
        "future_projection": impact_future_projection(text),
    }
    out["Impact"]["mean"] = sum(out["Impact"].values()) / len(out["Impact"].values())

    out["Commitment"] = {
        "statements": commitment_statements(text),
        "power": commitment_power(text),
    }
    out["Commitment"]["mean"] = sum(out["Commitment"].values()) / len(out["Commitment"].values())

    out["Scarcity/Urgency"] = {
        "temporal_urgency": scarcity_temporal_urgency(text),
        "exclusivity_quantity": scarcity_exclusivity_quantity(text),
    }
    out["Scarcity/Urgency"]["mean"] = sum(out["Scarcity/Urgency"].values()) / len(out["Scarcity/Urgency"].values())

    out["Engagement"] = {
        "identification": engagement_identification_pov(text),
        "self_reference": engagement_self_referencing(text),
        "inquiry": engagement_inquiry(text),
        # "narrative_transport": engagement_narrative_transport(text),
        "past": engagement_past(text),
        "imagery": engagement_imagery(text),
        "characters": engagement_characters(text),
    }
    out["Engagement"]["mean"] = sum(out["Engagement"].values()) / len(out["Engagement"].values())

    out["Propaganda"] = {
        "emotional_charge": prop_emotional_charge(text),
        "logical_distortion": prop_logical_distortion(text),
        "heuristic_identity_appeals": prop_heuristic_identity_appeals(text),
    }
    out["Propaganda"]["mean"] = sum(out["Propaganda"].values()) / len(out["Propaganda"].values())

    out["Style"] = {
        "fluency": style_fluency(text),
        "length": style_length(text),
        "rhetorical_punctuation": style_rhetorical_punctuation(text),
    }
    out["Style"]["mean"] = sum(out["Style"].values()) / len(out["Style"].values())

    return out

# -----------------------------------------------------------------------------
# Optional: prewarm caches for persistent mode.
# Set PI_PREWARM=1 to load heavy resources at startup (once per process).
# -----------------------------------------------------------------------------
if os.environ.get("PI_PREWARM", "").strip() == "1":
    try:
        _load_lexicons()
        _load_liwc_dic()
        _load_concreteness_dic()
        _load_mwe_concreteness_dic()
        _load_nrc_vad()
        _get_vader()
        _get_nlp()
    except Exception:
        pass